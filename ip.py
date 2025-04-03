import openpyxl
import subprocess
import socket
import requests
import ipaddress
import urllib3
import os 

# Подавляем предупреждения для HTTPS
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def ping_ip(ip):
    try:
        result = subprocess.run(['ping', '-c', '1', '-W', '1', ip],
                                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print('ping: ok')
        return result.returncode == 0
        
    except Exception:
        return False

def check_http(ip):
    try:
        response = requests.get(f"http://{ip}", timeout=3)
        print('http: ok')
        return response.status_code in [200, 301, 300]
        
    except Exception:
        return False

def check_https(ip):
    try:
        response = requests.get(f"https://{ip}", timeout=3, verify=False)
        print('https: ok')
        return response.status_code in [200, 301, 300]
         
    except Exception:
        return False

def check_ports(ip, ports):
    for port in ports:
        try:
            with socket.create_connection((ip, port), timeout=3):
                print('port: ok')
                return port
        except (socket.timeout, ConnectionRefusedError, OSError):
            continue
    return None

def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

def get_row_color(sheet, row_num):
    """
    Возвращает цвет строки, если он указан в стиле ячейки.
    """
    row_color = None
    for cell in sheet[row_num]:
        fill = cell.fill
        if fill and fill.start_color and fill.start_color.index:
            color = fill.start_color.index
            if color != "00000000":  # Если цвет задан (не прозрачный)
                row_color = color
                break
    return row_color

def list_sheets(file_path):
    """
    Возвращает список всех листов в Excel-файле.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return workbook.sheetnames
    except Exception as e:
        print(f"Ошибка при чтении Excel файла: {e}")
        return []

def find_excel_files():
    """
    Находит все .xlsx файлы в текущей директории.
    """
    current_directory = os.getcwd()
    files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
    return files

def main(file_path, output_path, log_file, sheet_name):
    try:
        # Загружаем Excel-файл с указанием имени листа
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except Exception as e:
        print(f"Ошибка при чтении Excel файла или листа: {e}")
        return

    ports_to_check = [22] + [56722] + list(range(55561, 55570))

    with open(log_file, 'w') as log:
        for row_num in range(2, sheet.max_row + 1):  # Пропускаем заголовки
            ip_cell = sheet.cell(row=row_num, column=1)  # Предполагаем, что IP-адрес в первой колонке
            ip = ip_cell.value
            row_color = get_row_color(sheet, row_num)

            if isinstance(ip, str) and validate_ip(ip):
                print(f"Проверяем IP: {ip}...  color  {row_color}")
                
                # Выполняем проверки
                is_pingable = ping_ip(ip)
                is_http = check_http(ip)
                is_https = check_https(ip)
                open_port = check_ports(ip, ports_to_check)

                is_accessible = is_pingable or is_http or is_https or open_port

                # Проверяем соответствие на основе цвета строки
                if row_color == 12 or row_color == 21:  # Желтый
                    print('ok')
                    if not is_accessible:
                        result = f"Несоответствие: IP {ip} не доступен, хотя в таблице занят."
                        print(result)
                        log.write(result + '\n')
                        # Логируем результаты проверок
                        log.write(f"Проверяем IP: {ip}...\n")
                        log.write(f"  Результаты:\n")
                        log.write(f"    Ping: {'Доступен' if is_pingable else 'Не доступен'}\n")
                        log.write(f"    HTTP: {'Доступен' if is_http else 'Не доступен'}\n")
                        log.write(f"    HTTPS: {'Доступен' if is_https else 'Не доступен'}\n")
                        log.write(f"    Открытый порт: {open_port if open_port else 'Нет доступных портов'}\n")

                elif row_color == 14:  # Зеленый
                    print('ok')
                    if is_accessible:
                        result = f"Несоответствие: IP {ip} доступен, хотя в таблице свободен."
                        print(result)
                        # Логируем результаты проверок
                        log.write(f"Проверяем IP: {ip}...\n")
                        log.write(f"  Результаты:\n")
                        log.write(f"    Ping: {'Доступен' if is_pingable else 'Не доступен'}\n")
                        log.write(f"    HTTP: {'Доступен' if is_http else 'Не доступен'}\n")
                        log.write(f"    HTTPS: {'Доступен' if is_https else 'Не доступен'}\n")
                        log.write(f"    Открытый порт: {open_port if open_port else 'Нет доступных портов'}\n")
                        log.write(result + '\n')
                        
            else:
                skip_msg = f"Пропускаем некорректный адрес: {ip}"
                print(skip_msg)
                log.write(skip_msg + '\n')

        # Сохраняем изменения в Excel-файле
        workbook.save(output_path)
        print(f"Результаты проверки сохранены в файл: {output_path}")
        print(f"Лог сохранён в файл: {log_file}")

if __name__ == "__main__":
    # Находим все .xlsx файлы в текущей директории
    excel_files = find_excel_files()
    if not excel_files:
        print("В текущей директории нет файлов с расширением .xlsx.")
        exit(1)

    # Выводим список найденных файлов
    print("Найденные .xlsx файлы:")
    for i, file in enumerate(excel_files, start=1):
        print(f"{i}. {file}")

    # Предлагаем выбрать файл
    try:
        choice = int(input("Введите номер файла для обработки: "))
        if 1 <= choice <= len(excel_files):
            selected_file = excel_files[choice - 1]
            print(f"Выбран файл: {selected_file}")
        else:
            print("Неверный выбор. Завершение программы.")
            exit(1)
    except ValueError:
        print("Некорректный ввод. Завершение программы.")
        exit(1)

    input_file = selected_file
    output_file = "results.xlsx"  # Укажите путь для сохранения результатов

    # Получаем список листов
    sheets = list_sheets(input_file)
    if not sheets:
        print("В файле нет доступных листов.")
        exit(1)

    # Выводим список листов
    print("Доступные листы:")
    for i, sheet in enumerate(sheets, start=1):
        print(f"{i}. {sheet}")

    # Предлагаем выбрать лист
    try:
        choice = int(input("Введите номер листа для сканирования: "))
        if 1 <= choice <= len(sheets):
            selected_sheet = sheets[choice - 1]
            print(f"Выбран лист: {selected_sheet}")
        else:
            print("Неверный выбор. Завершение программы.")
            exit(1)
    except ValueError:
        print("Некорректный ввод. Завершение программы.")
        exit(1)

    log_file = f"log{selected_sheet}.txt"
    # Запускаем основную программу с выбранным листом
    main(input_file, output_file, log_file, selected_sheet)